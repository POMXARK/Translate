#!/usr/bin/python
# -*- coding: UTF-8 -*-
# (90 * 8) + (40*8) + (9*4) + 37 + (3*8) + 70  = 1200 (600 ч или 25 полных дней ил же 75 рабочих дней по 8 часов)
#10^20
#10^3v0
#11^20
#11^50
#13 25
#45
#16 20

# cli.py
import click


import openpyxl
#import xlwt
#import xlrd
import winsound
import sys
import os
import os.path
import random
from msvcrt import getch
key = getch()
import time
import re
import requests
import json
from googletrans import Translator
translator = Translator()
###from translate import Translator
from difflib import SequenceMatcher
import re


global N_shogest_list
global  N_shogest_list_max
global col_bukv
global N_predlog
global n_english


@click.command()

def main():

    N_perevod = 0

    n_english = 1

    stop = 0


    wb = openpyxl.Workbook()
    ws = wb.active  # Add a sheet

    wb_p = openpyxl.Workbook()
    ws_p = wb_p.active  # Add a sheet

    good_par = []

    to_lang = 'English'
    secret = 'trnsl.1.1.20170910T021345Z.baf67aa28ccee9bc.2ff30aaf19e8930044683c365a3e5fbfeeeba79fSims_translate'

    print(" ИМЯ СОЗДОВАЕМОГО ФАЙЛА (БЕЗ РАСШИРЕНИЯ!) : ")
    text_edit = input()
    text_pairs =  text_edit + "_pairs.txt"
    print("ИМЯ ФАЙЛА С ОДИНОЧНЫМИ СЛОВАМИ : " , text_pairs)
    text_edit = text_edit + ".txt"
    print("ИМЯ ФАЙЛА : " , text_edit)
    print("ВВЕСТИ ТЕКСТ ИЛИ ВЫБРАТЬ ИЗ ФАЙЛА(t \ f или т \ ф)")
    sposob = input()

    if sposob == 't' or sposob == 'т':
        print("ИМЯ ФАЙЛА С ИСХОДНЫМ ТЕКСТОМ : N_" , text_edit)
        print("ВВЕЕДИТЕ ТЕКСТ")
        open("text0.txt", 'w' ,  encoding='utf-8')
     
        while True:
          text_save = input()
          if text_save:
            with open("text0.txt", 'a' ,  encoding='utf-8')as g:
                 g.write(text_save)
                 g.write("\n")

            with open("N_" + text_edit , 'a' ,  encoding='utf-8')as g:
                 g.write(text_save)
                 g.write("\n")             
          else:
            break    

        text = open("text0.txt" , encoding='utf-8')
            
    if sposob == 'f' or sposob == 'ф':
        print(" ИМЯ ФАЙЛА : ")
        faile = input()
        text = open(faile , encoding='utf-8')

    text_read = text.read()

    text_read = re.sub(r'[ ][(]', "(" , text_read )
    text_read = re.sub(r'[()]', "." , text_read )

    #text_read = re.sub(r'[•][ ]', "." , text_read )
    #text_read = re.sub(r' - ', "." , text_read )

    text_read = re.sub(r'\n', "." , text_read )
    text_read = re.sub(r'\n', "." , text_read )
    text_read = re.sub(r'\. ', "." , text_read )
    text_read = re.sub(r'\? ', "?" , text_read )
    text_read = re.sub(r'\! ', "!" , text_read )

    text_read = re.sub(r'\.  ', "." , text_read )
    text_read = re.sub(r'\?  ', "?" , text_read )
    text_read = re.sub(r'\!  ', "!" , text_read )

    text_read = re.sub(r'\.   ', "." , text_read )
    text_read = re.sub(r'\?   ', "?" , text_read )
    text_read = re.sub(r'\!   ', "!" , text_read )


    text_read = re.sub(r'\n', "." , text_read )
    text_read = re.sub(r'\. ', "." , text_read )
    text_read = re.sub(r'\? ', "?" , text_read )
    text_read = re.sub(r'\! ', "!" , text_read )

    text_read = re.sub(r'\.  ', "." , text_read )
    text_read = re.sub(r'\?  ', "?" , text_read )
    text_read = re.sub(r'\!  ', "!" , text_read )

    text_read = re.sub(r'\.   ', "." , text_read )
    text_read = re.sub(r'\?   ', "?" , text_read )
    text_read = re.sub(r'\!   ', "!" , text_read )


    text_read = re.sub(r'\.\.', "." , text_read )
    text_read = re.sub(r'\.\.\.', "." , text_read )
    text_read = re.sub(r'\.\.', "." , text_read )
    text_read = re.sub(r'\.\.\.', "." , text_read )


    #text_read = re.sub(r'([0-9][ ])', "." , text_read )
    #text_read = re.sub(r'([0-9])', "." , text_read )



    text_read = re.sub(r'([.A-Za-z][:][ ][A-Z])', "." , text_read )


    text_list = re.findall('(.*?(?:[.?!]|.$))', text_read )

    N = 1
    big_words = []

    time_sleep = [2,3,4]

    tr_eng = []



    #shogest_list_ru = []

    flag_pairs = 1

    last_ru = 0

    print("> ПОДОЖДИТЕ... >") 

    #for element in range(len(text_list)):
        #print("stop ",stop)
        #print("n_english :::",n_english)
        #if n_english == 0 or stop == 1:
            #print("n_english ",n_english)
            #print("ТОЧКА ЗАВЕРШЕНИЯ в начале!")
            #break


        
        
    if N > 0:
        ru = re.compile("[а-яА-Я]+")
        words_ru = text_list
        russian = [w for w in filter(ru.match, text_list)]
        print(russian)
        eng = re.compile("[a-zA-Z]+")
        words_eng = text_list
        english = [w for w in filter(eng.match, text_list)]
        print(english)
        n_english = len(english)



    ##ЭТАП ПЕРЕВОДА:
    for element_eng  in english:
        time.sleep(3)
        translation_all = translator.translate(element_eng , dest='ru')
        translation = translation_all.text
        tr_eng.append(translation)
        #insert((english.index(element_eng)),translation)
        #print(element_eng)

    #print("tr_eng ",tr_eng)
    print("Предложения переведены!")

    winsound.Beep(200,600)
    winsound.Beep(230,600)
    winsound.Beep(200,600)


    ##ЭТАП СРАВНЕНИЯ:
    for element_ru in russian:
        #print("<NEN>")
        
        mix_l = []
        max_l = []
        N_max_l = []
        N_min_l  = []
        flag = 1
        shogest_list = []
        #print(russian)
        #print(element_ru)
        for element_tr in tr_eng:
            raznica = []
            l_element_tr = len(element_tr)
            l_element_ru = len(element_ru)
            raznica.append(l_element_tr)
            raznica.append(l_element_ru)
            #print("element_tr ",element_tr)
            #print("element_ru ",element_ru)
            #print("l_element_tr ",l_element_tr)
            #print("l_element_ru ",l_element_ru)
            mix_l.append(l_element_tr)
            max_l.append(l_element_ru)

            N_max_raznica = (max(raznica))
            N_min_raznica = (min(raznica))
            
            N_max_l = (max(max_l))
            N_min_l = (min(mix_l))

            N_max_l_1 = N_max_l / 100
            size_l = N_min_l - (N_max_l*N_max_l_1)
            size_2 = (N_max_l/2)
            size_3 = (N_max_l/18) # 3

            max_raznica = N_max_raznica - N_min_raznica

            #print("raznica " , raznica)
            #print("max_raznica " , max_raznica)
            #print("N_min_l " ,N_min_l)
            #print("N_max_l ",N_max_l)
            #print("size_l " ,size_l)
            #print("size_l " ,size_l)
            #print("size_2 " ,size_2)
            #print("size_3 " ,size_3)
            
            ###ОТСЕЕВАНИЕ ПРЕДЛОЖЕНИЙ НЕ ПОХОЖИХ ПО ДЛИННЕ!
            
            if max_raznica < (N_max_raznica/3):
            #if ((N_min_l > size_2 and l_element_ru < 30) or (N_min_l > size_3 and l_element_ru >= 30) or (N_max_l >= 60 and N_min_l > 30)):
                #print("ПРЕДЛОЖЕНИЯ ПОДХОДЯТ! : ", element_tr, ' - ' , element_ru)            
                s = SequenceMatcher(lambda x: x==" ", element_tr , element_ru )
                shogest_0 = s.ratio()
                #print("схожесть ", shogest_0)
                #shogest_list.insert((russian.index(element_ru)),shogest_0)
                shogest_list.append(shogest_0)
                N_shogest_list_max = (max(shogest_list))
                N_shogest_list = shogest_list.index(N_shogest_list_max)
            else:
                #print("ПРЕДЛОЖЕНИЯ НЕ ПОДХОДЯТ! : ", element_tr, ' - ' , element_ru) 
                shogest_list.append(0)
                N_shogest_list_max = (max(shogest_list))
                N_shogest_list = shogest_list.index(N_shogest_list_max)
                                
            l_shogest_list = len(shogest_list)
            #print("l_shogest_list ",l_shogest_list)

            #print("N_shogest_list_max ",N_shogest_list_max)
                #print("N_shogest_list ",N_shogest_list)
                #print("english[N_shogest_list] ",english[N_shogest_list])
                #print("shogest_list ",shogest_list)
                #print("element_tr ",element_tr)

        while flag == 1:

            #print("N_shogest_list_max !",N_shogest_list_max)
            #print("N_shogest_list !",N_shogest_list)
            #print("english !",english)
            if N_shogest_list_max > 0.10:
                english[N_shogest_list].capitalize()
                element_ru = element_ru.capitalize()
                print("ПАРА?(y/n или д/н (Предложение выбранно не верно? жми v или в)) : " , english[N_shogest_list] ," - " , element_ru)


                variant = input()

            if (variant == 'v' or variant == 'в') and l_shogest_list > 0 :
                #print("shogest_list ",shogest_list)
                print("Подбор следующего предложения...")
                #print("l_shogest_list" ,l_shogest_list)
                for n, i in enumerate(shogest_list):

                    if i == N_shogest_list_max :

                        shogest_list[n] = 0
                        l_shogest_list -= 1
                        
                        
                N_shogest_list_max = (max(shogest_list))
                N_shogest_list = shogest_list.index(N_shogest_list_max)
                
                flag = 1
                
                continue




            if variant == 'n' or variant == 'н':
                print("ПОДОЖДИТЕ...")
                flag = 0
                flag_perebor = 0
                flag_pairs = 0
                break
            if variant == 'y' or variant == 'д':
                flag_pairs = 1
                with open ( text_edit, 'a' , encoding='utf-8') as g:

                    g.write(english[N_shogest_list])
                    g.write("#")
                    g.write(element_ru)
                    g.write("\n")
                    print("!!!ЗАПИСАННО!!!")
                    print("ПОДОЖДИТЕ...")
                    flag = 0


                N = 0;


                col_bukv = len(english[N_shogest_list])
                col_bukv_all = col_bukv
                col_bukv_1 = col_bukv / 100
                big_word = col_bukv / 40


                big_words = []
                if col_bukv > 100:
                    col_stop  = col_bukv_1 * 2                       
                if col_bukv > 80:
                    col_stop  = col_bukv_1 * 3
                elif col_bukv > 50:
                    col_stop  = col_bukv_1 * 4
                elif col_bukv < 50:
                    col_stop  = col_bukv_1 * 20                              
                elif col_bukv < 30:
                    col_stop  = col_bukv_1 * 60                



            while col_bukv > (col_stop):
                if flag_pairs == 0:
                    
                    for n, i in enumerate(english):

                        if i == english[N_shogest_list] :

                            english[n] = "0"


                    print("ПАРЫ СЛОВ ЗАПИСАННЫ!")
                    break

                if col_bukv < (col_stop):

                    break


                two_words = english[N_shogest_list].split(" ")
                for element in two_words:
                    len_element = len(element)
                    if len_element > big_word:
                        big_words.append(element)
                                           
                for element in big_words:
                                        #print("ТЫ ТУТ2333")
                    if flag_pairs == 0:
                                            #print("ПАРЫ СЛОВ ЗАПИСАННЫ!")
                        break
                    if col_bukv < (col_stop):
                        print("!!!ДОСТАТОЧНО!!! : ", col_stop)
                        print("ПОДОЖДИТЕ...")
                        flag = 0
                        break


                    print("ВВЕДИТЕ НУЖНОЕ КОЛИЧЕСТВО ПАР : ")
                    flag_pairs = int(input())

                    while flag_pairs > 0:
                        #print("ТЫ ТУТ23")
                        secure_random = random.SystemRandom()                            
                        random_word = secure_random.choice(big_words)
                        print("big_words :: " ,big_words)



                        random_word = re.sub(r'\?', "" , random_word )
                        random_word = re.sub(r' \?', "" , random_word )
                        random_word = re.sub(r'  \?', "" , random_word )
                        random_word = re.sub(r'\!', "" , random_word )
                        random_word = re.sub(r' \!', "" , random_word )
                        random_word = re.sub(r'  \!', "" , random_word )                            
                        random_word = re.sub(r'\.', "" , random_word )
                        random_word = re.sub(r'\  .', "" , random_word )
                        random_word = re.sub(r'\,', "" , random_word )
                        random_word = re.sub(r'\  ,', "" , random_word )
                        random_word = re.sub(r' / \.,', "" , random_word )

                        random_word = random_word.capitalize()

                        
                        time.sleep(1)
                        translation_words = translator.translate(random_word , dest='ru')
                        translation_w = translation_words.text
                        translation_w = translation_w.capitalize()
                        print("ПРЕДЛОЖЕНИЕ : " , english[N_shogest_list] ," - " , element_ru)


                        if random_word in good_par and len(big_words) > 1:
                            print('УЖЕ БЫЛО!!' ,random_word, " : ", good_par)
                            print("ПОДОЖДИТЕ...")
                            break                        


                        print("ПАРА СЛОВ?(y/n или д/н (для продолжения напишите  END. ) : " , random_word , " - " , translation_w)
                        variant_2 = input()

                        for n, i in enumerate(big_words):

                            i_r = i.capitalize()
                            i_r = re.sub(r'\?', "" , i_r )
                            i_r = re.sub(r' \?', "" , i_r )
                            i_r = re.sub(r'  \?', "" , i_r )
                            i_r = re.sub(r'\!', "" , i_r )
                            i_r = re.sub(r' \!', "" , i_r )
                            i_r = re.sub(r'  \!', "" , i_r )                            
                            i_r = re.sub(r'\.', "" , i_r )
                            i_r = re.sub(r'\  .', "" , i_r )
                            i_r = re.sub(r'\,', "" , i_r )
                            i_r = re.sub(r'\  ,', "" , i_r )
                            i_r = re.sub(r' / \.,', "" , i_r )

                            if i_r == random_word:

                                big_words.pop(n)


                            

                        if variant_2 == 'y' or variant_2 == 'д':
                            good_par.append(random_word)
                            flag_pairs -= 1
                            #print("flag_pairs :" ,flag_pairs)
                            print("ВЕРНЫЙ ВАРИАНТ ПЕРЕВОДА?(y/n или д/н) : ")
                            variant_3 = input()

                            if len(big_words) < 1:
                                flag_pairs = 0
                            
                            if variant_3 == 'y' or variant_3 == 'д':

                                with open ( text_pairs, 'a' , encoding='utf-8') as g:
                                    g.write(random_word)
                                    g.write("#")
                                    g.write(translation_w)
                                    g.write("\n")
            
                            if variant_3 == 'n' or variant_3 == 'н':

                                print("ВВЕДИТЕ ВЕРНЫЙ ВАРИАНТ ПЕРЕВОДА : ")
                                perevod = input()
                                with open ( text_pairs, 'a' , encoding='utf-8') as g:
                                    g.write(random_word)
                                    g.write("#")
                                    g.write(perevod)
                                    g.write("\n")

                            print("### ПАРА СЛОВ ЗАПИСАННА ###")
                            print("ПОДОЖДИТЕ...")

                            col_words = len(two_words)

                            col_bukv = col_bukv - (col_bukv_1 * (10*(col_bukv_all/col_words)))


                            print("***ЗАПИСАННО***")
                            print("ПОДОЖДИТЕ...")
                            flag = 0
                        if variant_2 == 'n':
                            if len(big_words)== 0:
                                flag = 0
                                print("ЗАКОНЧИЛИСЬ СЛОВА!")
                                break
                            print("***ПРОПУСК***")
                            print("ПОДОЖДИТЕ...")
                            flag = 0
                            
                            continue
                            
           # flag = 0
           # break

       # n_english -=1
        #break
            
                                         
    #print("tr_eng ",tr_eng)
        


    f = open(text_edit, 'r+' , encoding='utf-8')
    data = f.readlines()
    for i in range(len(data)):
        row = data[i].split('#')
        for j in range(len(row)):
            col_i = i + 1
            col_j = j + 1
            #print("i,j,row[j] :" ,col_i ," " ,col_j ," " ,row[j])
            ws.cell(row=col_i,column=col_j).value = row[j] # Write to cell i, j
        text_edit_exel = text_edit.replace('.txt', '')
        wb.save(text_edit_exel + '.xlsx')
        f.close()


    f_p = open(text_pairs, 'r+' , encoding='utf-8')
    data_p = f_p.readlines()
    for i_p in range(len(data_p)):
        row_p = data_p[i_p].split('#')
        for j_p in range(len(row_p)):
            col_i_p = i_p + 1
            col_j_p = j_p + 1
            #print("col_i_p,col_j_p,row_p[j_p] :" ,col_i_p ," " ,col_j_p ," " ,row_p[j_p])
            ws_p.cell(row=col_i_p,column=col_j_p).value = row_p[j_p] # Write to cell i, j
        text_pairs_exel = text_pairs.replace('.txt', '')
        wb_p.save(text_pairs_exel + '.xlsx')
        f_p.close()


if __name__ == "__main__":
    main()
